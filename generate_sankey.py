#!/usr/bin/env python3
"""
Transaction Sankey Dashboard Generator
Usage: python3 generate_sankey.py your_transactions.csv
Output: transactions.xlsx + Sankey_Chart.html (in same folder as input file)
"""

import sys
import os
import json
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ── CONFIG — edit these if your columns change ────────────────────────────────
DATE_COL        = 'Date'
MERCHANT_COL    = 'Merchant Name'
DESCRIPTION_COL = 'Description'
AMOUNT_COL      = 'Amount'
CATEGORY_COL    = 'Category'
INCOME_CATEGORY = 'Income'          # Category name that represents income
EXCLUDE_CATS    = ['Internal Transfers']  # Categories to ignore entirely
# ─────────────────────────────────────────────────────────────────────────────

DARK_BG    = "1A1A2E"; ACCENT1  = "16213E"; ACCENT2 = "0F3460"
GOLD       = "E94560";  TEAL    = "00B4D8"; GREEN   = "06D6A0"
PURPLE     = "7B2D8B";  WHITE   = "FFFFFF"; LIGHT_GRAY = "F8F9FA"
MID_GRAY   = "DEE2E6";  TEXT_DARK = "212529"

CAT_COLORS = {
    'Bills':'#E63946','Charity':'#06D6A0','Cash':'#ADB5BD','Eating Out':'#FF6B35',
    'Entertainment':'#7B2D8B','Finances':'#4895EF','General':'#ADB5BD',
    'Groceries':'#52B788','Health & Beauty':'#F72585','Home & Family':'#7209B7',
    'Investment':'#4CC9F0','Shopping':'#F77F00','Transport':'#3A0CA3','Travel':'#560BAD',
}
MONTH_NAMES = {1:'Jan',2:'Feb',3:'Mar',4:'Apr',5:'May',6:'Jun',
               7:'Jul',8:'Aug',9:'Sep',10:'Oct',11:'Nov',12:'Dec'}

def hdr(cell, bg=ACCENT2, fg=WHITE, size=11, bold=True):
    cell.font = Font(name='Arial', bold=bold, color=fg, size=size)
    cell.fill = PatternFill('solid', start_color=bg)
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = Border(bottom=Side(style='thin', color=MID_GRAY),
                         right=Side(style='thin', color=MID_GRAY))

def dat(cell, bg=WHITE, bold=False, right=False, fmt=None):
    cell.font = Font(name='Arial', size=10, bold=bold, color=TEXT_DARK)
    cell.fill = PatternFill('solid', start_color=bg)
    cell.alignment = Alignment(horizontal='right' if right else 'left', vertical='center')
    cell.border = Border(bottom=Side(style='hair', color=MID_GRAY),
                         right=Side(style='hair', color=MID_GRAY))
    if fmt: cell.number_format = fmt

def title_row(ws, text, cols, row=1, height=32):
    ws.merge_cells(f'A{row}:{get_column_letter(cols)}{row}')
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(name='Arial', bold=True, size=14, color=WHITE)
    c.fill = PatternFill('solid', start_color=DARK_BG)
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[row].height = height


def build(csv_path):
    if not os.path.exists(csv_path):
        print(f"ERROR: File not found: {csv_path}"); sys.exit(1)

    out_dir   = os.getcwd()
    xlsx_path = os.path.join(out_dir, 'transactions.xlsx')
    html_path = os.path.join(out_dir, 'Sankey_Chart.html')

    print(f"Reading: {csv_path}")
    df = pd.read_csv(csv_path)
    df[DATE_COL] = pd.to_datetime(df[DATE_COL])
    df['Month']     = df[DATE_COL].dt.month
    df['Year']      = df[DATE_COL].dt.year
    df['MonthName'] = df[DATE_COL].dt.strftime('%b')
    df['MonthYear'] = df[DATE_COL].dt.strftime('%b %Y')

    years = sorted(df['Year'].unique())
    expense_cats = sorted([c for c in df[CATEGORY_COL].unique()
                           if c not in [INCOME_CATEGORY] + EXCLUDE_CATS])
    expense_df = df[(df[AMOUNT_COL] < 0) & (~df[CATEGORY_COL].isin(EXCLUDE_CATS))].copy()
    expense_df[AMOUNT_COL] = expense_df[AMOUNT_COL].abs()

    wb = openpyxl.Workbook()

    # ── Sheet 1: Transactions ──────────────────────────────────────────────────
    ws = wb.active; ws.title = "📋 Transactions"
    ws.sheet_view.showGridLines = False; ws.sheet_properties.tabColor = ACCENT2
    cols  = ['Date','Merchant','Description','Amount','Category','Month','Year','MonthYear']
    widths = [13, 22, 32, 12, 18, 8, 7, 12]
    title_row(ws, "💳  Transaction Data", len(cols))
    for i,(h,w) in enumerate(zip(cols,widths),1):
        hdr(ws.cell(row=2,column=i,value=h))
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[2].height = 24
    export = df[[DATE_COL,MERCHANT_COL,DESCRIPTION_COL,AMOUNT_COL,
                 CATEGORY_COL,'Month','Year','MonthYear']].copy()
    export[DATE_COL] = export[DATE_COL].dt.strftime('%Y-%m-%d')
    for ri, row in enumerate(export.itertuples(index=False), 3):
        bg = WHITE if ri%2==1 else LIGHT_GRAY
        for ci, val in enumerate(row, 1):
            c = ws.cell(row=ri, column=ci, value=val)
            dat(c, bg=bg, right=(ci==4), fmt=('£#,##0.00;[Red](£#,##0.00)' if ci==4 else None))
    ws.freeze_panes = 'A3'

    # ── Sheet 2: Monthly Summary ───────────────────────────────────────────────
    ws2 = wb.create_sheet("📊 Monthly Summary")
    ws2.sheet_view.showGridLines = False; ws2.sheet_properties.tabColor = TEAL
    s_hdrs  = ['Year','Month','Month Name','Income (£)','Expenses (£)','Net (£)'] + expense_cats
    s_widths = [8,8,12,14,14,14] + [14]*len(expense_cats)
    title_row(ws2, "📊  Monthly Income vs Expenses", len(s_hdrs))
    for i,(h,w) in enumerate(zip(s_hdrs,s_widths),1):
        hdr(ws2.cell(row=2,column=i,value=h))
        ws2.column_dimensions[get_column_letter(i)].width = w
    ws2.row_dimensions[2].height = 24
    ri = 3
    for (y,m), grp in df.groupby(['Year','Month']):
        inc = grp[(grp[CATEGORY_COL]==INCOME_CATEGORY)&(grp[AMOUNT_COL]>0)][AMOUNT_COL].sum()
        exp = abs(grp[(grp[AMOUNT_COL]<0)&(~grp[CATEGORY_COL].isin(EXCLUDE_CATS))][AMOUNT_COL].sum())
        net = inc - exp
        bg  = WHITE if ri%2==1 else LIGHT_GRAY
        vals = [int(y), int(m), MONTH_NAMES[m], inc, exp, net]
        for cat in expense_cats:
            vals.append(abs(grp[(grp[AMOUNT_COL]<0)&(grp[CATEGORY_COL]==cat)][AMOUNT_COL].sum()))
        for ci, val in enumerate(vals, 1):
            c = ws2.cell(row=ri, column=ci, value=val)
            dat(c, bg=bg, right=(ci>=4), fmt=('£#,##0.00' if ci>=4 else None))
            if ci==6: c.font = Font(name='Arial',size=10,bold=True,
                                    color=("00A86B" if val>=0 else "E63946"))
        ri += 1
    ws2.freeze_panes = 'A3'

    # ── Sheet 3: Categories ────────────────────────────────────────────────────
    ws3 = wb.create_sheet("🏷️ Categories")
    ws3.sheet_view.showGridLines = False; ws3.sheet_properties.tabColor = GREEN
    title_row(ws3, "🏷️  Spending by Category (All Time)", 5)
    c_hdrs  = ['Category','Total Spent (£)','Transactions','Avg per Txn (£)','% of Spending']
    c_widths = [20,16,14,16,14]
    for i,(h,w) in enumerate(zip(c_hdrs,c_widths),1):
        hdr(ws3.cell(row=2,column=i,value=h))
        ws3.column_dimensions[get_column_letter(i)].width = w
    ws3.row_dimensions[2].height = 24
    cat_df = expense_df.groupby(CATEGORY_COL)[AMOUNT_COL].agg(['sum','count']).reset_index()
    cat_df.columns = [CATEGORY_COL,'Total','Count']
    cat_df['Avg'] = cat_df['Total']/cat_df['Count']
    cat_df['Pct'] = cat_df['Total']/cat_df['Total'].sum()
    cat_df = cat_df.sort_values('Total',ascending=False)
    for ri, row in enumerate(cat_df.itertuples(index=False), 3):
        bg = WHITE if ri%2==1 else LIGHT_GRAY
        cc = CAT_COLORS.get(getattr(row,CATEGORY_COL), MID_GRAY).lstrip('#')
        c1 = ws3.cell(row=ri, column=1, value=getattr(row,CATEGORY_COL))
        c1.font = Font(name='Arial',size=10,bold=True,color=WHITE)
        c1.fill = PatternFill('solid',start_color=cc)
        c1.alignment = Alignment(horizontal='left',vertical='center',indent=1)
        c1.border = Border(bottom=Side(style='hair',color=MID_GRAY))
        for ci,(val,fmt) in enumerate(zip([row.Total,row.Count,row.Avg,row.Pct],
                                          ['£#,##0.00','#,##0','£#,##0.00','0.0%']),2):
            c = ws3.cell(row=ri,column=ci,value=val)
            dat(c,bg=bg,right=True,fmt=fmt)
    ws3.freeze_panes = 'A3'

    # ── Sheet 4: Sankey Data ───────────────────────────────────────────────────
    ws4 = wb.create_sheet("🌊 Sankey Data")
    ws4.sheet_view.showGridLines = False; ws4.sheet_properties.tabColor = GOLD
    title_row(ws4, "🌊  Sankey Flow Data — open Sankey_Chart.html in browser to view", 6)
    ws4['A3'] = "Filter Year:"
    ws4['A3'].font = Font(name='Arial',bold=True,size=11,color=WHITE)
    ws4['A3'].fill = PatternFill('solid',start_color=ACCENT2)
    ws4['A3'].alignment = Alignment(horizontal='right',vertical='center')
    ws4['B3'] = "All"
    ws4['B3'].font = Font(name='Arial',bold=True,size=11,color=DARK_BG)
    ws4['B3'].fill = PatternFill('solid',start_color=TEAL)
    ws4['B3'].alignment = Alignment(horizontal='center',vertical='center')
    year_dv = DataValidation(type="list",formula1=f'"All,{",".join(str(y) for y in years)}"',showDropDown=False)
    ws4.add_data_validation(year_dv); year_dv.add(ws4['B3'])
    ws4['C3'] = "Filter Month:"
    ws4['C3'].font = Font(name='Arial',bold=True,size=11,color=WHITE)
    ws4['C3'].fill = PatternFill('solid',start_color=ACCENT2)
    ws4['C3'].alignment = Alignment(horizontal='right',vertical='center')
    ws4['D3'] = "All"
    ws4['D3'].font = Font(name='Arial',bold=True,size=11,color=DARK_BG)
    ws4['D3'].fill = PatternFill('solid',start_color=TEAL)
    ws4['D3'].alignment = Alignment(horizontal='center',vertical='center')
    month_dv = DataValidation(type="list",formula1='"All,Jan,Feb,Mar,Apr,May,Jun,Jul,Aug,Sep,Oct,Nov,Dec"',showDropDown=False)
    ws4.add_data_validation(month_dv); month_dv.add(ws4['D3'])
    sk_hdrs  = ['From (Source)','To (Category)','Amount (£)','Year','Month','MonthYear']
    sk_widths = [18,20,14,8,8,12]
    for i,(h,w) in enumerate(zip(sk_hdrs,sk_widths),1):
        hdr(ws4.cell(row=5,column=i,value=h))
        ws4.column_dimensions[get_column_letter(i)].width = w
    ws4.row_dimensions[5].height = 22
    sk_rows = []
    for (y,m), grp in df.groupby(['Year','Month']):
        inc = grp[(grp[CATEGORY_COL]==INCOME_CATEGORY)&(grp[AMOUNT_COL]>0)][AMOUNT_COL].sum()
        for cat in expense_cats:
            exp = abs(grp[(grp[AMOUNT_COL]<0)&(grp[CATEGORY_COL]==cat)][AMOUNT_COL].sum())
            if exp > 0:
                sk_rows.append({'from':'Income','to':cat,'amount':exp,
                                'year':int(y),'month':MONTH_NAMES[m],
                                'monthYear':f"{MONTH_NAMES[m]} {y}",
                                'income_ref': float(inc)})
    for ri, row in enumerate(sk_rows, 6):
        bg = WHITE if ri%2==0 else LIGHT_GRAY
        for ci,val in enumerate([row['from'],row['to'],row['amount'],
                                  row['year'],row['month'],row['monthYear']],1):
            c = ws4.cell(row=ri,column=ci,value=val)
            dat(c,bg=bg,right=(ci==3),fmt=('£#,##0.00' if ci==3 else None))
    ws4.freeze_panes = 'A6'

    # ── Sheet 5: How to Use ────────────────────────────────────────────────────
    ws5 = wb.create_sheet("ℹ️ How to Use")
    ws5.sheet_view.showGridLines = False; ws5.sheet_properties.tabColor = PURPLE
    ws5.column_dimensions['A'].width = 2; ws5.column_dimensions['B'].width = 60
    ws5.merge_cells('A1:B1')
    tc = ws5['A1']
    tc.value = "ℹ️  How to Use Your Transaction Sankey Dashboard"
    tc.font = Font(name='Arial',bold=True,size=16,color=WHITE)
    tc.fill = PatternFill('solid',start_color=DARK_BG)
    tc.alignment = Alignment(horizontal='center',vertical='center',wrap_text=True)
    ws5.row_dimensions[1].height = 40
    steps = [
        ("",""),("📋  TRANSACTIONS SHEET",""),
        ("→","All raw transaction data. Use Data → AutoFilter to filter any column."),
        ("",""),("📊  MONTHLY SUMMARY SHEET",""),
        ("→","Month-by-month: Total Income, Expenses, Net, and each spending category."),
        ("→","Green Net = you saved. Red Net = you overspent."),
        ("",""),("🏷️  CATEGORIES SHEET",""),
        ("→","All-time spending ranked by category with totals and % of spending."),
        ("",""),("🌊  SANKEY CHART",""),
        ("→","Open Sankey_Chart.html in Chrome/Firefox for the interactive Sankey."),
        ("→","Use Year and Month dropdowns to filter. Hover flows for exact amounts."),
        ("→","Wider flow = more money going to that category."),
        ("",""),("🔄  REGENERATING",""),
        ("→","Run: python3 generate_sankey.py your_new_file.csv"),
        ("→","Columns expected: Date (A), Merchant (B), Description (C), Amount (D), Category (E)."),
    ]
    for ri,(b,t) in enumerate(steps,2):
        ws5.row_dimensions[ri].height = 22
        is_section = b.startswith(("📋","📊","🏷️","🌊","🔄","ℹ️"))
        bc = ws5.cell(row=ri,column=1,value=b)
        tc2 = ws5.cell(row=ri,column=2,value=t)
        if is_section:
            for c in [bc,tc2]:
                c.font = Font(name='Arial',bold=True,size=12,color=WHITE)
                c.fill = PatternFill('solid',start_color=ACCENT2)
            ws5.row_dimensions[ri].height = 26
        elif t:
            bc.font = Font(name='Arial',size=10,color=TEAL,bold=True)
            tc2.font = Font(name='Arial',size=10,color=TEXT_DARK)
            for c in [bc,tc2]: c.fill = PatternFill('solid',start_color=LIGHT_GRAY)
            tc2.alignment = Alignment(vertical='center',wrap_text=True)

    wb.move_sheet("ℹ️ How to Use", offset=-4)
    wb.save(xlsx_path)
    print(f"✅  Excel saved:  {xlsx_path}")

    # ── HTML Sankey ────────────────────────────────────────────────────────────
    js_data = '[\n' + ',\n'.join(json.dumps(r) for r in sk_rows) + '\n]'
    cat_colors_js = json.dumps(CAT_COLORS)

    # Build flat transactions list for drill-down (exclude internal transfers)
    txn_rows = []
    for idx, (_, row) in enumerate(df[~df[CATEGORY_COL].isin(EXCLUDE_CATS)].iterrows()):
        if row[AMOUNT_COL] >= 0: continue  # expenses only
        txn_rows.append({
            'id':  idx,
            'date':  row[DATE_COL].strftime('%Y-%m-%d'),
            'merchant': str(row[MERCHANT_COL]) if pd.notna(row[MERCHANT_COL]) else '',
            'desc':  str(row[DESCRIPTION_COL]) if pd.notna(row[DESCRIPTION_COL]) else '',
            'amount': round(abs(float(row[AMOUNT_COL])), 2),
            'cat':   str(row[CATEGORY_COL]),
            'year':  int(row['Year']),
            'month': str(row['MonthName']),
        })
    txns_js = '[\n' + ',\n'.join(json.dumps(r) for r in txn_rows) + '\n]'
    html = """<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8"><meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Transaction Sankey</title>
<style>
* { margin:0; padding:0; box-sizing:border-box; }
body { background:#1A1A2E; color:#fff; font-family:Arial,sans-serif; }
header { background:#0F3460; padding:20px 32px; display:flex; align-items:center; gap:16px; border-bottom:2px solid #E94560; }
header h1 { font-size:22px; font-weight:700; }
#pl { font-size:13px; color:#90CAF9; margin-left:auto; }
.controls { background:#16213E; padding:16px 32px; display:flex; gap:16px; align-items:center; flex-wrap:wrap; border-bottom:1px solid #0F3460; }
.controls label { font-size:13px; color:#90CAF9; font-weight:600; }
.controls select { background:#0F3460; color:#fff; border:1px solid #E94560; border-radius:6px; padding:8px 14px; font-size:13px; cursor:pointer; outline:none; }
.stats { background:#16213E; padding:12px 32px; display:flex; gap:24px; flex-wrap:wrap; border-bottom:1px solid #0F3460; }
.sc { background:#0F3460; border-radius:8px; padding:12px 20px; min-width:140px; border-left:3px solid #E94560; }
.sc.inc { border-left-color:#06D6A0; } .sc.exp { border-left-color:#E94560; }
.sc.np  { border-left-color:#4CC9F0; } .sc.nn  { border-left-color:#FF6B35; }
.sl { font-size:11px; color:#90CAF9; text-transform:uppercase; letter-spacing:.5px; }
.sv { font-size:20px; font-weight:700; margin-top:4px; }
.sv.iv { color:#06D6A0; } .sv.ev { color:#E94560; } .sv.pv { color:#4CC9F0; } .sv.nv { color:#FF6B35; } .sv.cv { color:#c084fc; }
.cat-panel { background:#16213E; padding:14px 32px; border-bottom:1px solid #0F3460; }
.cat-panel-header { display:flex; align-items:center; gap:16px; margin-bottom:10px; }
.cat-panel-header span { font-size:13px; color:#90CAF9; font-weight:600; }
.cat-toggle-btn { background:none; border:1px solid #0F3460; color:#90CAF9; border-radius:5px; padding:4px 10px; font-size:11px; cursor:pointer; }
.cat-toggle-btn:hover { border-color:#E94560; color:#fff; }
.cat-chips { display:flex; flex-wrap:wrap; gap:8px; }
.chip { display:flex; align-items:center; gap:6px; padding:5px 12px 5px 8px; border-radius:20px; font-size:12px; font-weight:600; cursor:pointer; border:2px solid transparent; user-select:none; transition: opacity .15s, border-color .15s; }
.chip .cdot { width:10px; height:10px; border-radius:50%; flex-shrink:0; }
.chip.off { opacity:0.35; border-color:transparent; }
.chip.on { opacity:1; border-color:rgba(255,255,255,0.25); }
#cc { padding:24px 32px; overflow-x:auto; }
#sg { background:#16213E; border-radius:12px; border:1px solid #0F3460; display:block; min-width:600px; }
.tip { position:fixed; background:rgba(15,52,96,.97); border:1px solid #E94560; border-radius:8px; padding:10px 16px; font-size:13px; pointer-events:none; display:none; z-index:999; max-width:240px; box-shadow:0 4px 20px rgba(0,0,0,.5); }
.tip strong { color:#E94560; display:block; margin-bottom:4px; } .tip span { color:#e0e0e0; }
.leg { padding:0 32px 16px; display:flex; flex-wrap:wrap; gap:10px; }
.li { display:flex; align-items:center; gap:6px; background:#16213E; border-radius:6px; padding:6px 12px; font-size:12px; }
.ld { width:12px; height:12px; border-radius:3px; flex-shrink:0; }
#txn-panel { margin:0 32px 40px; border-radius:12px; overflow:hidden; border:1px solid #0F3460; display:none; }
#txn-header { display:flex; align-items:center; justify-content:space-between; padding:14px 20px; background:#0F3460; }
#txn-title { font-size:15px; font-weight:700; }
#txn-title span { font-size:12px; font-weight:400; color:#90CAF9; margin-left:10px; }
#txn-close { background:none; border:1px solid #E94560; color:#E94560; border-radius:5px; padding:4px 12px; font-size:12px; cursor:pointer; }
#txn-close:hover { background:#E94560; color:#fff; }
#txn-search-bar { background:#16213E; padding:10px 20px; border-bottom:1px solid #0F3460; }
#txn-search { background:#0F3460; border:1px solid #0F3460; color:#fff; border-radius:6px; padding:7px 14px; font-size:13px; width:280px; outline:none; }
#txn-search:focus { border-color:#E94560; }
#txn-search::placeholder { color:#4a6fa5; }
#txn-table-wrap { overflow-x:auto; max-height:420px; overflow-y:auto; }
table { width:100%; border-collapse:collapse; font-size:13px; }
thead th { background:#0a2440; color:#90CAF9; font-size:11px; text-transform:uppercase; letter-spacing:.5px; padding:10px 16px; text-align:left; position:sticky; top:0; z-index:2; }
thead th.num { text-align:right; }
tbody tr { border-bottom:1px solid #0F3460; transition:background .1s; }
tbody tr:hover { background:#1e3a5f; }
tbody td { padding:10px 16px; color:#e0e0e0; }
tbody td.num { text-align:right; color:#E94560; font-weight:600; font-variant-numeric:tabular-nums; }
tbody td.date { color:#90CAF9; white-space:nowrap; }
tbody td.merchant { font-weight:600; color:#fff; max-width:160px; overflow:hidden; text-overflow:ellipsis; white-space:nowrap; }
tbody td.desc { color:#aab4c8; max-width:240px; overflow:hidden; text-overflow:ellipsis; white-space:nowrap; }
.cat-badge { display:inline-block; padding:2px 8px; border-radius:10px; font-size:11px; font-weight:600; color:#fff; }
#txn-footer { background:#0a2440; padding:10px 20px; font-size:12px; color:#90CAF9; display:flex; gap:24px; }
#txn-footer strong { color:#fff; }
</style>
</head>
<body>
<header><span>&#127754;</span><h1>Income Flow &mdash; Transaction Sankey</h1><span id="pl">All Time</span></header>
<div class="controls">
  <label>Year:</label><select id="ys"><option value="all">All Years</option></select>
  <label>Month:</label>
  <select id="ms">
    <option value="all">All Months</option>
    <option>Jan</option><option>Feb</option><option>Mar</option><option>Apr</option>
    <option>May</option><option>Jun</option><option>Jul</option><option>Aug</option>
    <option>Sep</option><option>Oct</option><option>Nov</option><option>Dec</option>
  </select>
</div>
<div class="cat-panel">
  <div class="cat-panel-header">
    <span>&#127991; Categories</span>
    <button class="cat-toggle-btn" id="btn-all">Select All</button>
    <button class="cat-toggle-btn" id="btn-none">Deselect All</button>
  </div>
  <div class="cat-chips" id="chips"></div>
</div>
<div class="stats">
  <div class="sc inc"><div class="sl">Total Income</div><div class="sv iv" id="si">-</div></div>
  <div class="sc exp"><div class="sl">Total Spent</div><div class="sv ev" id="se">-</div></div>
  <div class="sc" id="nc"><div class="sl">Net Savings</div><div class="sv" id="sn">-</div></div>
  <div class="sc" style="border-left-color:#7B2D8B"><div class="sl">Coverage</div><div class="sv cv" id="scov">-</div></div>
</div>
<div id="cc"><svg id="sg"></svg></div>
<div class="leg" id="leg"></div>
<div id="txn-panel">
  <div id="txn-header">
    <div id="txn-title">Transactions</div>
    <button id="txn-close">&#10005; Close</button>
  </div>
  <div id="txn-search-bar" style="display:flex;align-items:center;gap:12px;flex-wrap:wrap;">
    <input id="txn-search" type="text" placeholder="Search merchant or description...">
    <button class="cat-toggle-btn" id="txn-sel-all">Select All</button>
    <button class="cat-toggle-btn" id="txn-sel-none">Deselect All</button>
    <span id="txn-sel-hint" style="font-size:11px;color:#4a6fa5;margin-left:4px;">Deselected transactions are excluded from the chart</span>
  </div>
  <div id="txn-table-wrap">
    <table>
      <thead><tr>
        <th style="width:36px;text-align:center;"><input type="checkbox" id="chk-all" checked title="Select all visible"></th>
        <th>Date</th><th>Merchant</th><th>Description</th><th class="num">Amount</th>
      </tr></thead>
      <tbody id="txn-body"></tbody>
    </table>
  </div>
  <div id="txn-footer">
    <div>Showing <strong id="txn-count">0</strong> transactions</div>
    <div>Selected total: <strong id="txn-total">-</strong></div>
    <div id="txn-excl-msg" style="color:#E94560;display:none;">&#9888; Some transactions excluded from chart</div>
  </div>
</div>
<div class="tip" id="tip"></div>
<script>
var RAW  = __RAW_DATA__;
var CC   = __CAT_COLORS__;
var TXNS = __TXNS__;
var IC   = '#06D6A0';

// Set of excluded transaction IDs (excluded = deselected by user)
var excludedTxns = {};

function fmt(v) {
  return '\u00a3' + v.toLocaleString('en-GB', {minimumFractionDigits:2, maximumFractionDigits:2});
}

// ── Category chips ──────────────────────────────────────────────────────────
var allCats = [];
RAW.forEach(function(r) { if (allCats.indexOf(r.to) === -1) allCats.push(r.to); });
allCats.sort();
var enabledCats = {};
allCats.forEach(function(c) { enabledCats[c] = true; });

var chipsEl = document.getElementById('chips');
allCats.forEach(function(cat) {
  var col = CC[cat] || '#90CAF9';
  var chip = document.createElement('div');
  chip.className = 'chip on'; chip.dataset.cat = cat;
  chip.style.background = col + '22';
  chip.innerHTML = '<div class="cdot" style="background:' + col + '"></div><span>' + cat + '</span>';
  chip.addEventListener('click', function() {
    enabledCats[cat] = !enabledCats[cat];
    chip.className = 'chip ' + (enabledCats[cat] ? 'on' : 'off');
    render();
  });
  chipsEl.appendChild(chip);
});
document.getElementById('btn-all').addEventListener('click', function() {
  allCats.forEach(function(c) { enabledCats[c] = true; });
  document.querySelectorAll('.chip').forEach(function(ch) { ch.className = 'chip on'; });
  render();
});
document.getElementById('btn-none').addEventListener('click', function() {
  allCats.forEach(function(c) { enabledCats[c] = false; });
  document.querySelectorAll('.chip').forEach(function(ch) { ch.className = 'chip off'; });
  render();
});

// ── Year/month dropdowns ────────────────────────────────────────────────────
var yrs = [];
RAW.forEach(function(r) { if (yrs.indexOf(r.year) === -1) yrs.push(r.year); });
yrs.sort();
var ys = document.getElementById('ys');
yrs.forEach(function(y) { var o = document.createElement('option'); o.value = y; o.textContent = y; ys.appendChild(o); });
ys.addEventListener('change', render);
document.getElementById('ms').addEventListener('change', render);

// ── Transaction drill-down ──────────────────────────────────────────────────
var activeCat = null, activeYear = 'all', activeMonth = 'all';
var visibleTxnIds = [];  // IDs currently shown in the table

function showTxns(cat) {
  activeCat = cat;
  document.getElementById('txn-search').value = '';
  filterAndRenderTxns();
  var panel = document.getElementById('txn-panel');
  panel.style.display = 'block';
  panel.scrollIntoView({behavior:'smooth', block:'start'});
}

document.getElementById('txn-close').addEventListener('click', function() {
  document.getElementById('txn-panel').style.display = 'none';
  activeCat = null;
});

document.getElementById('txn-search').addEventListener('input', filterAndRenderTxns);

// Select/deselect all visible rows
document.getElementById('txn-sel-all').addEventListener('click', function() {
  visibleTxnIds.forEach(function(id) { delete excludedTxns[id]; });
  filterAndRenderTxns();
  render();
});
document.getElementById('txn-sel-none').addEventListener('click', function() {
  visibleTxnIds.forEach(function(id) { excludedTxns[id] = true; });
  filterAndRenderTxns();
  render();
});

// Header checkbox: select/deselect all visible
document.getElementById('chk-all').addEventListener('change', function() {
  if (this.checked) {
    visibleTxnIds.forEach(function(id) { delete excludedTxns[id]; });
  } else {
    visibleTxnIds.forEach(function(id) { excludedTxns[id] = true; });
  }
  filterAndRenderTxns();
  render();
});

function filterAndRenderTxns() {
  var search = document.getElementById('txn-search').value.toLowerCase();
  var rows = TXNS.filter(function(t) {
    return t.cat === activeCat
      && (activeYear === 'all' || String(t.year) === String(activeYear))
      && (activeMonth === 'all' || t.month === activeMonth)
      && (!search || t.merchant.toLowerCase().indexOf(search) > -1 || t.desc.toLowerCase().indexOf(search) > -1);
  });
  rows.sort(function(a,b) { return a.date < b.date ? 1 : -1; });

  visibleTxnIds = rows.map(function(t) { return t.id; });

  var col = CC[activeCat] || '#90CAF9';
  document.getElementById('txn-title').innerHTML = activeCat + ' <span>' + rows.length + ' transactions</span>';

  var allSelected = rows.every(function(t) { return !excludedTxns[t.id]; });
  var noneSelected = rows.every(function(t) { return excludedTxns[t.id]; });
  var chkAll = document.getElementById('chk-all');
  chkAll.checked = allSelected;
  chkAll.indeterminate = !allSelected && !noneSelected;

  var tbody = document.getElementById('txn-body');
  tbody.innerHTML = '';
  rows.forEach(function(t) {
    var enabled = !excludedTxns[t.id];
    var tr = document.createElement('tr');
    tr.style.opacity = enabled ? '1' : '0.38';
    var chkTd = document.createElement('td');
    chkTd.style.textAlign = 'center';
    var chk = document.createElement('input');
    chk.type = 'checkbox';
    chk.checked = enabled;
    chk.style.cursor = 'pointer';
    (function(txn, row) {
      chk.addEventListener('change', function() {
        if (this.checked) { delete excludedTxns[txn.id]; row.style.opacity = '1'; }
        else              { excludedTxns[txn.id] = true;  row.style.opacity = '0.38'; }
        // update header checkbox state
        var allSel = visibleTxnIds.every(function(id) { return !excludedTxns[id]; });
        var noneSel = visibleTxnIds.every(function(id) { return excludedTxns[id]; });
        chkAll.checked = allSel;
        chkAll.indeterminate = !allSel && !noneSel;
        updateTxnFooter();
        render();
      });
    })(t, tr);
    chkTd.appendChild(chk);
    tr.appendChild(chkTd);
    function makeTd(cls, text) {
      var td = document.createElement('td');
      if (cls) td.className = cls;
      td.textContent = text;
      return td;
    }
    tr.appendChild(makeTd('date', t.date));
    tr.appendChild(makeTd('merchant', t.merchant || '-'));
    tr.appendChild(makeTd('desc', t.desc || '-'));
    tr.appendChild(makeTd('num', fmt(t.amount)));
    tbody.appendChild(tr);
  });

  updateTxnFooter();
}

function updateTxnFooter() {
  var rows = TXNS.filter(function(t) {
    return t.cat === activeCat
      && (activeYear === 'all' || String(t.year) === String(activeYear))
      && (activeMonth === 'all' || t.month === activeMonth);
  });
  var selectedRows = rows.filter(function(t) { return !excludedTxns[t.id]; });
  var total = selectedRows.reduce(function(s,t){return s+t.amount;}, 0);
  document.getElementById('txn-count').textContent = selectedRows.length + ' / ' + rows.length;
  document.getElementById('txn-total').textContent = fmt(total);
  var hasExcluded = Object.keys(excludedTxns).length > 0;
  document.getElementById('txn-excl-msg').style.display = hasExcluded ? 'block' : 'none';
}

// ── Main render ─────────────────────────────────────────────────────────────
function render() {
  activeYear  = ys.value;
  activeMonth = document.getElementById('ms').value;
  document.getElementById('pl').textContent = (activeYear === 'all' ? 'All Years' : activeYear) + (activeMonth === 'all' ? '' : ', ' + activeMonth);

  // Build category map from TXNS directly (respects per-transaction exclusions)
  var cm = {}, ip = {};
  TXNS.forEach(function(t) {
    if (!enabledCats[t.cat]) return;
    if (excludedTxns[t.id]) return;
    if (activeYear !== 'all' && String(t.year) !== String(activeYear)) return;
    if (activeMonth !== 'all' && t.month !== activeMonth) return;
    cm[t.cat] = (cm[t.cat] || 0) + t.amount;
  });

  // Income comes from RAW (period totals, unaffected by txn exclusions)
  var fPeriod = RAW.filter(function(r) {
    return (activeYear === 'all' || String(r.year) === String(activeYear)) && (activeMonth === 'all' || r.month === activeMonth);
  });
  fPeriod.forEach(function(r) { var k = r.year + '-' + r.month; if (!ip[k]) ip[k] = r.income_ref; });
  var ti = Object.values(ip).reduce(function(a,b){return a+b;}, 0);
  var te = Object.values(cm).reduce(function(a,b){return a+b;}, 0);
  var net = ti - te;
  document.getElementById('si').textContent = fmt(ti);
  document.getElementById('se').textContent = fmt(te);
  var sn = document.getElementById('sn');
  sn.textContent = (net >= 0 ? '+' : '') + fmt(Math.abs(net));
  sn.className = 'sv ' + (net >= 0 ? 'pv' : 'nv');
  document.getElementById('nc').className = 'sc ' + (net >= 0 ? 'np' : 'nn');
  document.getElementById('scov').textContent = ti > 0 ? (te/ti*100).toFixed(1) + '%' : '-';
  draw(cm, ti, te);

  var leg = document.getElementById('leg');
  leg.innerHTML = '';
  var ent = Object.entries(cm).filter(function(e) { return e[1] > 0; }).sort(function(a,b) { return b[1]-a[1]; });
  ent.forEach(function(e) {
    var cat = e[0];
    var div = document.createElement('div'); div.className = 'li';
    var dot = document.createElement('div'); dot.className = 'ld'; dot.style.background = CC[cat] || '#90CAF9';
    var sp  = document.createElement('span'); sp.textContent = cat;
    div.appendChild(dot); div.appendChild(sp); leg.appendChild(div);
  });
  if (activeCat) filterAndRenderTxns();
}

// ── SVG helpers ─────────────────────────────────────────────────────────────
function mkEl(tag, attrs) {
  var el = document.createElementNS('http://www.w3.org/2000/svg', tag);
  Object.keys(attrs).forEach(function(k) { el.setAttribute(k, attrs[k]); });
  return el;
}

function draw(cm, ti, te) {
  var svg = document.getElementById('sg');
  svg.innerHTML = '';
  var ent = Object.entries(cm).filter(function(e) { return e[1] > 0; }).sort(function(a,b) { return b[1]-a[1]; });
  if (!ent.length) {
    svg.setAttribute('width', '100%'); svg.setAttribute('height', '200');
    var t = mkEl('text', {x:'50%', y:'100', 'text-anchor':'middle', fill:'#90CAF9', 'font-size':'18', 'font-family':'Arial'});
    t.textContent = 'No data for selected period / categories';
    svg.appendChild(t);
    return;
  }

  var NODE_MIN_H = 60, GAP = 12, PAD = 40;
  var tot = ent.reduce(function(s,e){return s+e[1];}, 0);
  var rawH = ent.map(function(e) { return Math.max(NODE_MIN_H, (e[1]/tot) * Math.max(NODE_MIN_H*ent.length, 600)); });
  var totalCatH = rawH.reduce(function(a,b){return a+b;},0) + GAP*(ent.length-1);
  var H = totalCatH + PAD*2;
  var W = Math.max(600, svg.parentElement.clientWidth - 48);
  svg.setAttribute('width', W); svg.setAttribute('height', H);
  svg.setAttribute('viewBox', '0 0 ' + W + ' ' + H);

  var nW=28, left=130, right=W-160, usH=H-PAD*2;
  var iH = Math.min(usH*0.88, totalCatH), iY = PAD + (usH-iH)/2;
  var cY = PAD + (usH-totalCatH)/2;

  var nodes = ent.map(function(e, i) {
    var n = {cat:e[0], val:e[1], y:cY, h:rawH[i]};
    cY += rawH[i] + GAP;
    return n;
  });

  var tip = document.getElementById('tip');
  function showTip(e, html) { tip.innerHTML = html; tip.style.display = 'block'; moveTip(e); }
  function moveTip(e) { tip.style.left = (e.clientX+14)+'px'; tip.style.top = (e.clientY-10)+'px'; }
  function hideTip() { tip.style.display = 'none'; }

  // Flows
  var incOffset = 0;
  nodes.forEach(function(n) {
    var col = CC[n.cat] || '#90CAF9';
    var frac = n.val / (ti || tot);
    var fH = Math.max(3, frac * iH);
    var srcY = iY + incOffset;
    incOffset += fH;
    var x0=left+nW, x1=right, mx=(x0+x1)/2;
    var d = 'M'+x0+','+srcY+' C'+mx+','+srcY+' '+mx+','+n.y+' '+x1+','+n.y
          + ' L'+x1+','+(n.y+n.h)+' C'+mx+','+(n.y+n.h)+' '+mx+','+(srcY+fH)+' '+x0+','+(srcY+fH)+' Z';
    var path = mkEl('path', {d:d, fill:col, opacity:'0.4'});
    path.style.cursor = 'pointer'; path.style.transition = 'opacity .2s';
    (function(node) {
      path.addEventListener('mouseenter', function(e) {
        path.setAttribute('opacity','0.72');
        showTip(e, '<strong>'+node.cat+'</strong><span>'+fmt(node.val)+'<br>'+(node.val/(ti||1)*100).toFixed(1)+'% of income<br>'+(node.val/(tot||1)*100).toFixed(1)+'% of spending<br><em style="color:#4CC9F0">Click to see transactions</em></span>');
      });
      path.addEventListener('mousemove', moveTip);
      path.addEventListener('mouseleave', function() { path.setAttribute('opacity','0.4'); hideTip(); });
      path.addEventListener('click', function() { hideTip(); showTxns(node.cat); });
    })(n);
    svg.appendChild(path);
  });

  // Income node
  svg.appendChild(mkEl('rect', {x:left, y:iY, width:nW, height:iH, rx:6, fill:IC}));
  function addTxt(x, y, txt, fill, size, bold, anchor) {
    var t = mkEl('text', {x:x, y:y, 'text-anchor':anchor||'end', 'dominant-baseline':'middle', fill:fill, 'font-size':size, 'font-weight':bold?'bold':'normal', 'font-family':'Arial'});
    t.textContent = txt; svg.appendChild(t);
  }
  addTxt(left-10, iY+iH/2-10, 'Income', '#06D6A0', 14, true, 'end');
  addTxt(left-10, iY+iH/2+10, fmt(ti), '#90CAF9', 11, false, 'end');

  // Category nodes
  nodes.forEach(function(n) {
    var col = CC[n.cat] || '#90CAF9';
    var r = mkEl('rect', {x:right, y:n.y, width:nW, height:n.h, rx:4, fill:col});
    r.style.cursor = 'pointer';
    (function(node) {
      r.addEventListener('mouseenter', function(e) {
        showTip(e, '<strong>'+node.cat+'</strong><span>'+fmt(node.val)+'<br>'+(node.val/(tot||1)*100).toFixed(1)+'% of spending<br><em style="color:#4CC9F0">Click to see transactions</em></span>');
      });
      r.addEventListener('mousemove', moveTip);
      r.addEventListener('mouseleave', hideTip);
      r.addEventListener('click', function() { hideTip(); showTxns(node.cat); });
    })(n);
    svg.appendChild(r);
    var lbl = mkEl('text', {x:right+nW+10, y:n.y+n.h/2-8, 'text-anchor':'start', 'dominant-baseline':'middle', fill:'#e0e0e0', 'font-size':'12', 'font-weight':'bold', 'font-family':'Arial'});
    lbl.style.cursor = 'pointer';
    lbl.textContent = n.cat;
    (function(node) { lbl.addEventListener('click', function() { showTxns(node.cat); }); })(n);
    svg.appendChild(lbl);
    var amt = mkEl('text', {x:right+nW+10, y:n.y+n.h/2+9, 'text-anchor':'start', 'dominant-baseline':'middle', fill:'#90CAF9', 'font-size':'11', 'font-family':'Arial'});
    amt.textContent = fmt(n.val);
    svg.appendChild(amt);
  });
}

render();
window.addEventListener('resize', render);
</script>
</body>
</html>"""
    html = html.replace('__RAW_DATA__', js_data).replace('__CAT_COLORS__', cat_colors_js).replace('__TXNS__', txns_js)

    with open(html_path, 'w') as f:
        f.write(html)
    print(f"✅  HTML saved:   {html_path}")
    print(f"\n📊  Summary:")
    print(f"    Rows processed : {len(df):,}")
    print(f"    Sankey flows   : {len(sk_rows)}")
    print(f"    Date range     : {df[DATE_COL].min().date()} → {df[DATE_COL].max().date()}")
    print(f"    Years covered  : {', '.join(str(y) for y in years)}")
    print(f"\nDone! Open Sankey_Chart.html in your browser.")


if __name__ == '__main__':
    if len(sys.argv) < 2:
        print("Usage: python3 generate_sankey.py your_transactions.csv")
        sys.exit(1)
    build(sys.argv[1])
